"""
Exports comptables CP/RTT à une date donnée.

Génère un classeur Excel avec 2 onglets :
- Synthèse CP (jours)
- Synthèse RTT (heures)
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from models import db
from models.conge import Conge
from models.parametrage import AllocationConge
from models.user import User


def _style_header_xlsx(ws, row=1):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="008C3A", end_color="008C3A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(ws, min_width=12, max_width=42):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def export_compta_cp_rtt_xlsx(parametrage, as_of: date, include_inactifs: bool = False) -> io.BytesIO:
    """
    Exporte une synthèse CP/RTT à une date donnée, sur l'exercice du parametrage.

    Règles de consommation :
    - CP = somme des jours ouvrables des congés validés type CP + Anciennete
    - RTT = somme des heures RTT des congés validés type RTT
    - pris en compte si date_debut >= debut_exercice ET date_fin <= as_of
    """
    if parametrage is None:
        raise ValueError("Paramétrage annuel manquant.")

    debut = parametrage.debut_exercice
    fin = parametrage.fin_exercice
    as_of_clamped = min(max(as_of, debut), fin)

    users_q = User.query
    if not include_inactifs:
        users_q = users_q.filter_by(actif=True)
    users = users_q.order_by(User.nom, User.prenom).all()
    user_ids = [u.id for u in users]

    allocations = {}
    if user_ids:
        alloc_rows = AllocationConge.query.filter(
            AllocationConge.parametrage_id == parametrage.id,
            AllocationConge.user_id.in_(user_ids),
        ).all()
        allocations = {a.user_id: a for a in alloc_rows}

    cp_consumed = {uid: 0 for uid in user_ids}
    if user_ids:
        rows = (
            db.session.query(
                Conge.user_id,
                func.coalesce(func.sum(Conge.nb_jours_ouvrables), 0),
            )
            .filter(
                Conge.user_id.in_(user_ids),
                Conge.statut == "valide",
                Conge.type_conge.in_(["CP", "Anciennete"]),
                Conge.date_debut >= debut,
                Conge.date_fin <= as_of_clamped,
            )
            .group_by(Conge.user_id)
            .all()
        )
        cp_consumed.update({uid: int(val or 0) for uid, val in rows})

    rtt_consumed = {uid: 0 for uid in user_ids}
    if user_ids:
        rows = (
            db.session.query(
                Conge.user_id,
                func.coalesce(func.sum(Conge.nb_heures_rtt), 0),
            )
            .filter(
                Conge.user_id.in_(user_ids),
                Conge.statut == "valide",
                Conge.type_conge == "RTT",
                Conge.date_debut >= debut,
                Conge.date_fin <= as_of_clamped,
            )
            .group_by(Conge.user_id)
            .all()
        )
        rtt_consumed.update({uid: int(val or 0) for uid, val in rows})

    wb = Workbook()
    ws_cp = wb.active
    ws_cp.title = "Synthèse CP"
    ws_rtt = wb.create_sheet("Synthèse RTT")

    title = (
        f"Export comptable au {as_of_clamped.strftime('%d/%m/%Y')} "
        f"(exercice {debut.strftime('%d/%m/%Y')} → {fin.strftime('%d/%m/%Y')})"
    )
    for ws in (ws_cp, ws_rtt):
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([])

    headers_cp = ["Salarié", "Alloué (jours)", "Consommé (jours)", "Reste (jours)", "Actif"]
    ws_cp.append(headers_cp)
    _style_header_xlsx(ws_cp, ws_cp.max_row)

    total_alloue_cp = 0
    total_consomme_cp = 0
    for u in users:
        a = allocations.get(u.id)
        alloue = int(getattr(a, "total_jours", 0) or 0)
        consomme = int(cp_consumed.get(u.id, 0) or 0)
        ws_cp.append([f"{u.prenom} {u.nom}", alloue, consomme, alloue - consomme, "Oui" if u.actif else "Non"])
        total_alloue_cp += alloue
        total_consomme_cp += consomme

    ws_cp.append([])
    ws_cp.append(["TOTAL", total_alloue_cp, total_consomme_cp, total_alloue_cp - total_consomme_cp, ""])
    ws_cp[ws_cp.max_row][0].font = Font(bold=True)

    headers_rtt = ["Salarié", "Alloué (heures)", "Consommé (heures)", "Reste (heures)", "Actif"]
    ws_rtt.append(headers_rtt)
    _style_header_xlsx(ws_rtt, ws_rtt.max_row)

    total_alloue_rtt = 0
    total_consomme_rtt = 0
    for u in users:
        a = allocations.get(u.id)
        alloue = int(getattr(a, "total_rtt_heures", 0) or 0)
        consomme = int(rtt_consumed.get(u.id, 0) or 0)
        ws_rtt.append([f"{u.prenom} {u.nom}", alloue, consomme, alloue - consomme, "Oui" if u.actif else "Non"])
        total_alloue_rtt += alloue
        total_consomme_rtt += consomme

    ws_rtt.append([])
    ws_rtt.append(["TOTAL", total_alloue_rtt, total_consomme_rtt, total_alloue_rtt - total_consomme_rtt, ""])
    ws_rtt[ws_rtt.max_row][0].font = Font(bold=True)

    _autosize_columns(ws_cp)
    _autosize_columns(ws_rtt)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


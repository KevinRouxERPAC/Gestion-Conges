"""Utilitaires de mise en forme partagés pour les exports Excel (openpyxl).

Centralise le style d'en-tête ERPAC et l'ajustement de largeur des colonnes, qui
étaient auparavant dupliqués à l'identique dans ``export.py``,
``export_comptable.py`` et ``export_interessement.py``.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Vert ERPAC, utilisé pour les en-têtes dans toute l'application.
HEADER_COLOR = "008C3A"


def style_header_xlsx(ws, row=1):
    """Applique le style d'en-tête ERPAC (fond vert, texte blanc gras, bordures
    fines, texte centré et renvoyé à la ligne) à toutes les cellules de la ligne
    ``row`` (1-based)."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, min_width=10, max_width=48):
    """Ajuste la largeur de chaque colonne à la longueur de son contenu, bornée
    par ``[min_width, max_width]``."""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))

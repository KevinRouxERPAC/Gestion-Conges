"""
Service d'export Excel et PDF pour les congés.
"""
# import_salaries: see services/import_salaries.py for sync logic

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


def _style_header_xlsx(ws, row=1):
    """Applique des styles aux en-têtes Excel. row: numéro de ligne (1-based)."""
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="008C3A", end_color="008C3A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.border = thin_border
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def export_conges_excel(conges, user_nom="", user_prenom=""):
    """
    Génère un fichier Excel des congés.
    conges: liste d'objets Conge
    Retourne: BytesIO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Congés"

    titre = f"Congés - {user_prenom} {user_nom}" if user_nom else "Congés"
    ws.append([titre])
    ws.merge_cells("A1:E1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Date début", "Date fin", "Jours ouvrables", "Type", "Statut"]
    ws.append(headers)
    _style_header_xlsx(ws, ws.max_row)

    for c in conges:
        ws.append([
            c.date_debut.strftime("%d/%m/%Y"),
            c.date_fin.strftime("%d/%m/%Y"),
            c.nb_jours_ouvrables,
            c.type_conge,
            getattr(c, "statut", "valide") or "valide",
        ])

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_conges_equipe_excel(users_with_conges):
    """
    Génère un fichier Excel de tous les congés par salarié.
    users_with_conges: liste de dicts {"user": User, "conges": [Conge]}
    Retourne: BytesIO
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Congés équipe"

    ws.append(["Export des congés - Équipe"])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Export du {date.today().strftime('%d/%m/%Y')}"])
    ws.append([])

    headers = ["Salarié", "Date début", "Date fin", "Jours", "Type", "Statut"]
    ws.append(headers)
    _style_header_xlsx(ws, ws.max_row)

    for item in users_with_conges:
        user = item["user"]
        conges = item.get("conges", [])
        nom_complet = f"{user.prenom} {user.nom}"
        for i, c in enumerate(conges):
            ws.append([
                nom_complet if i == 0 else "",
                c.date_debut.strftime("%d/%m/%Y"),
                c.date_fin.strftime("%d/%m/%Y"),
                c.nb_jours_ouvrables,
                c.type_conge,
                getattr(c, "statut", "valide") or "valide",
            ])

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_conges_pdf(conges, solde_info, user_nom="", user_prenom=""):
    """
    Génère un PDF récapitulatif des congés d'un salarié.
    Retourne: BytesIO
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    titre_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=12,
    )

    elements = []

    titre = f"Récapitulatif des congés - {user_prenom} {user_nom}" if user_nom else "Récapitulatif des congés"
    elements.append(Paragraph(titre, titre_style))
    elements.append(Paragraph(f"Export du {date.today().strftime('%d/%m/%Y')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    if solde_info:
        elements.append(Paragraph("<b>Solde</b>", styles["Heading2"]))
        solde_data = [
            ["Total alloué", str(solde_info.get("total_alloue", 0))],
            ["Total consommé", str(solde_info.get("total_consomme", 0))],
            ["Solde restant", str(solde_info.get("solde_restant", 0))],
        ]
        t_solde = Table(solde_data, colWidths=[6 * cm, 4 * cm])
        t_solde.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f4f6")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ]))
        elements.append(t_solde)
        elements.append(Spacer(1, 0.5 * cm))

    elements.append(Paragraph("<b>Historique des congés</b>", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * cm))

    if conges:
        data = [["Période", "Jours", "Type", "Statut"]]
        for c in conges:
            data.append([
                f"{c.date_debut.strftime('%d/%m/%Y')} - {c.date_fin.strftime('%d/%m/%Y')}",
                str(c.nb_jours_ouvrables),
                c.type_conge,
                getattr(c, "statut", "valide") or "valide",
            ])
        t = Table(data, colWidths=[6 * cm, 3 * cm, 4 * cm, 3 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#008C3A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph("Aucun congé enregistré.", styles["Normal"]))

    elements.append(Spacer(1, 1 * cm))
    elements.append(Paragraph("— ERPAC Gestion des Congés —", ParagraphStyle(
        "footer", parent=styles["Normal"], fontSize=8, textColor=colors.grey, alignment=1
    )))

    doc.build(elements)
    buffer.seek(0)
    return buffer

import io



from openpyxl import Workbook

from openpyxl.styles import Font



from services.interessement import calculer_interessement

from services.export_utils import (

    autosize_columns as _autosize_columns,

    style_header_xlsx as _style_header_xlsx,

)





def export_interessement_xlsx(periode, include_inactifs: bool = False) -> io.BytesIO:

    res = calculer_interessement(periode, include_inactifs=include_inactifs)



    wb = Workbook()

    ws = wb.active

    ws.title = "Synthese"

    ws_d = wb.create_sheet("Detail")



    montant_total = float(periode.montant_total_euros or 0)

    malus = float(getattr(periode, "malus_maladie_par_jour", None) or 5.0)

    title = f"Interessement - {periode.libelle} ({periode.date_debut} au {periode.date_fin})"

    for w in (ws, ws_d):

        w.append([title])

        w.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        w["A1"].font = Font(bold=True, size=13)

        w.append([])



    if montant_total > 0:

        ws.append([f"Montant total a repartir : {montant_total:.2f} EUR"])

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)

        ws["A3"].font = Font(bold=True)

        ws.append([])



    ws.append([f"Malus maladie : {malus} points / jour (seul type impactant)"])

    ws.append([])



    headers = ["Salarie", "Actif", "Jours maladie", "Malus total", "Points final"]

    if montant_total > 0:

        headers.append("Montant (EUR)")

    ws.append(headers)

    _style_header_xlsx(ws, ws.max_row)



    total_euros = 0.0

    for r in res:

        row = [

            f"{r.prenom} {r.nom}",

            "Oui" if r.actif else "Non",

            r.jours_maladie,

            round(r.total_malus, 2),

            round(r.points_final, 2),

        ]

        if montant_total > 0:

            part = r.part_euros if r.part_euros is not None else 0.0

            total_euros += part

            row.append(round(part, 2))

        ws.append(row)



    if montant_total > 0 and res:

        total_row = ["TOTAL", "", "", "", round(sum(r.points_final for r in res), 2), round(total_euros, 2)]

        ws.append(total_row)

        for cell in ws[ws.max_row]:

            cell.font = Font(bold=True)



    headers_d = ["Salarie", "Type absence", "Jours", "Points/jour", "Impact"]

    ws_d.append(headers_d)

    _style_header_xlsx(ws_d, ws_d.max_row)



    for r in res:

        nom = f"{r.prenom} {r.nom}"

        if not r.details:

            ws_d.append([nom, "Maladie", 0, malus, 0])

            continue

        for i, d in enumerate(r.details):

            ws_d.append([

                nom if i == 0 else "",

                d.type_absence,

                d.jours,

                d.points_par_jour,

                round(d.impact_points, 2),

            ])



    _autosize_columns(ws)

    _autosize_columns(ws_d)



    buf = io.BytesIO()

    wb.save(buf)

    buf.seek(0)

    return buf


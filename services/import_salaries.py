"""
Import / mise a jour des salaries depuis CSV ou Excel.
Cle: identifiant. Creation si absent, mise a jour si present.
"""
import csv
import io
import re
from datetime import datetime
from typing import Callable

from openpyxl import load_workbook

# Format docs/current_usage (ex. CONGES SALARIES.csv) : NOM, PRENOM (identifiant optionnel)
COL_IDENTIFIANT = ("identifiant", "login", "matricule")
COL_NOM = ("nom", "nom_famille", "lastname", "nom")
COL_PRENOM = ("prenom", "prenom_usage", "firstname", "prenom")
COL_EMAIL = ("email", "mail", "courriel")
COL_DATE_EMBAUCHE = ("date_embauche", "date_entree", "embauche")
COL_ROLE = ("role", "type", "profil")
COL_ACTIF = ("actif", "actif_yn", "statut")


def _norm(s):
    if not s:
        return ""
    s = str(s).strip().lower().replace(" ", "_")
    return re.sub(r"_+", "_", s)


def _find_col(headers, candidates):
    for i, h in enumerate(headers):
        n = _norm(h)
        for c in candidates:
            if _norm(c) == n:
                return i
    return None


def _parse_date(val):
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if hasattr(val, "date"):
        return val.date()
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(val):
    if val is None:
        return True
    s = str(val).strip().lower()
    if s in ("1", "oui", "yes", "true", "o"):
        return True
    if s in ("0", "non", "no", "false", "n"):
        return False
    return True


def _normalize_ident(s):
    """Retourne une chaîne pour identifiant : minuscules, sans accents, alphanumerique + underscore."""
    if not s:
        return ""
    s = str(s).strip().lower()
    for old, new in (("e", "e"), ("e", "e"), ("a", "a"), ("u", "u"), ("i", "i"), ("o", "o"), ("c", "c")):
        s = s.replace(old, new)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or "user"


def _generate_identifiant(nom, prenom):
    return (_normalize_ident(prenom) + "_" + _normalize_ident(nom)) or "user"


def _ensure_identifiants_unique(rows):
    """Attribue un identifiant aux lignes qui n'en ont pas, puis rend la liste unique (suffixe _2, _3...)."""
    seen = {}
    for row in rows:
        ident = (row.get("identifiant") or "").strip()
        if not ident:
            ident = _generate_identifiant(row.get("nom") or "", row.get("prenom") or "")
            row["identifiant"] = ident
        key = ident
        if key in seen:
            seen[key] += 1
            row["identifiant"] = key + "_" + str(seen[key])
        else:
            seen[key] = 0
    return rows


def _find_header_row(all_rows):
    """Trouve la ligne d'en-tete parmi les 10 premieres (format docs: titre puis NOM;PRENOM;...)."""
    for idx in range(min(10, len(all_rows))):
        headers = [str(h).strip() for h in all_rows[idx]]
        if _find_col(headers, COL_NOM) is not None and _find_col(headers, COL_PRENOM) is not None:
            return idx, headers
    return None, []


def parse_csv(content):
    """Parse un CSV (format docs/current_usage: NOM;PRENOM;... ou identifiant;nom;prenom;...)."""
    text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
    delimiter = ";" if ";" in (text.split("\n")[0] if text else "") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)
    if not all_rows:
        return []
    header_idx, headers = _find_header_row(all_rows)
    if header_idx is None:
        return []
    first_row = headers
    idx_id = _find_col(headers, COL_IDENTIFIANT)
    idx_nom = _find_col(headers, COL_NOM)
    idx_prenom = _find_col(headers, COL_PRENOM)
    if idx_nom is None or idx_prenom is None:
        return []
    idx_email = _find_col(headers, COL_EMAIL)
    idx_date = _find_col(headers, COL_DATE_EMBAUCHE)
    idx_role = _find_col(headers, COL_ROLE)
    idx_actif = _find_col(headers, COL_ACTIF)

    rows = []
    for row in all_rows[header_idx + 1:]:
        row = list(row) if row else []
        max_idx = max(idx_nom, idx_prenom) if idx_id is None else max(idx_id, idx_nom, idx_prenom)
        if len(row) <= max_idx:
            continue
        identifiant = (row[idx_id] or "").strip() if idx_id is not None else ""
        nom = (row[idx_nom] or "").strip()
        prenom = (row[idx_prenom] or "").strip()
        if not nom or not prenom:
            continue
        email = (row[idx_email] or "").strip() if idx_email is not None and idx_email < len(row) else ""
        date_embauche = _parse_date(row[idx_date] if idx_date is not None and idx_date < len(row) else None)
        role_raw = (row[idx_role] or "salarie").strip() if idx_role is not None and idx_role < len(row) else "salarie"
        role = "rh" if str(role_raw).lower() in ("rh", "admin") else "salarie"
        actif = _parse_bool(row[idx_actif] if idx_actif is not None and idx_actif < len(row) else True)
        rows.append({
            "identifiant": identifiant,
            "nom": nom,
            "prenom": prenom,
            "email": email or None,
            "date_embauche": date_embauche,
            "role": role,
            "actif": actif,
        })
    return _ensure_identifiants_unique(rows)


def parse_excel(content):
    """Parse la premiere feuille Excel. Meme structure que parse_csv."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in first_row]
    idx_id = _find_col(headers, COL_IDENTIFIANT)
    idx_nom = _find_col(headers, COL_NOM)
    idx_prenom = _find_col(headers, COL_PRENOM)
    if idx_nom is None or idx_prenom is None:
        return []
    idx_email = _find_col(headers, COL_EMAIL)
    idx_date = _find_col(headers, COL_DATE_EMBAUCHE)
    idx_role = _find_col(headers, COL_ROLE)
    idx_actif = _find_col(headers, COL_ACTIF)

    rows = []
    for row in rows_iter:
        row = list(row) if row else []
        max_idx = max(idx_nom, idx_prenom) if idx_id is None else max(idx_id, idx_nom, idx_prenom)
        if len(row) <= max_idx:
            continue
        identifiant = (row[idx_id] or "").strip() if idx_id is not None and idx_id < len(row) else ""
        nom = (row[idx_nom] or "").strip() if idx_nom < len(row) else ""
        prenom = (row[idx_prenom] or "").strip() if idx_prenom < len(row) else ""
        if not nom or not prenom:
            continue
        email = (row[idx_email] or "").strip() if idx_email is not None and idx_email < len(row) else ""
        date_embauche = _parse_date(row[idx_date] if idx_date is not None and idx_date < len(row) else None)
        role_raw = (row[idx_role] or "salarie")
        role = "rh" if str(role_raw).lower() in ("rh", "admin") else "salarie"
        actif = _parse_bool(row[idx_actif] if idx_actif is not None and idx_actif < len(row) else True)
        rows.append({
            "identifiant": identifiant,
            "nom": nom,
            "prenom": prenom,
            "email": email or None,
            "date_embauche": date_embauche,
            "role": role,
            "actif": actif,
        })
    return _ensure_identifiants_unique(rows)


def sync_users(rows, default_password, hash_password):
    """
    Cree ou met a jour les utilisateurs. Cle: identifiant.
    Retourne: (created, updated, errors).
    """
    from models.user import User
    from models import db

    created = 0
    updated = 0
    errors = []

    for i, row in enumerate(rows):
        identifiant = (row.get("identifiant") or "").strip()
        nom = (row.get("nom") or "").strip()
        prenom = (row.get("prenom") or "").strip()
        if not nom or not prenom:
            errors.append("Ligne %d: nom et prenom obligatoires." % (i + 2))
            continue
        if not identifiant:
            identifiant = _generate_identifiant(nom, prenom)
        user = User.query.filter_by(identifiant=identifiant).first()
        if user:
            user.nom = nom
            user.prenom = prenom
            user.email = (row.get("email") or "").strip() or None
            user.date_embauche = row.get("date_embauche")
            user.role = (row.get("role") or "salarie").strip() or "salarie"
            if user.role != "rh":
                user.role = "salarie"
            user.actif = row.get("actif", True)
            updated += 1
        else:
            if not default_password or len(default_password) < 6:
                errors.append("Ligne %d (%s): mot de passe par defaut requis (min. 6 car.) pour les nouveaux." % (i + 2, identifiant))
                continue
            new_user = User(
                identifiant=identifiant,
                nom=nom,
                prenom=prenom,
                mot_de_passe_hash=hash_password(default_password),
                role=(row.get("role") or "salarie").strip() or "salarie",
                actif=row.get("actif", True),
                email=(row.get("email") or "").strip() or None,
                date_embauche=row.get("date_embauche"),
            )
            if new_user.role != "rh":
                new_user.role = "salarie"
            db.session.add(new_user)
            created += 1

    return created, updated, errors

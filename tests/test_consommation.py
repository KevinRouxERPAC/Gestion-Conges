"""Tests de la source de vérité unique du calcul de consommation (NFR9).

Vérifie que la primitive `somme_consommation` se comporte correctement et que
les différents consommateurs (solde, export comptable, intéressement) produisent
des chiffres cohérents pour un même salarié/exercice.
"""
from datetime import date

import openpyxl

from models.conge import Conge
from services.consommation import (
    somme_consommation,
    STATUT_VALIDE,
    STATUTS_EN_ATTENTE,
    TYPES_CP,
    TYPE_RTT,
)
from services.solde import calculer_jours_cps_consommes, calculer_heures_rtt_consommes
from services.calcul_jours import compter_jours_ouvrables_avec_demi
from services.export_comptable import export_compta_cp_rtt_xlsx
from services.interessement import calculer_interessement


def _ajouter_conge(db_session, user_id, **kwargs):
    defaults = dict(
        date_debut=date(2026, 6, 1),
        date_fin=date(2026, 6, 5),
        nb_jours_ouvrables=5,
        type_conge="CP",
        statut="valide",
    )
    defaults.update(kwargs)
    c = Conge(user_id=user_id, **defaults)
    db_session.session.add(c)
    db_session.session.commit()
    return c


class TestPrimitive:
    def test_scalaire_valide(self, db_session, users, parametrage):
        _ajouter_conge(db_session, users["salarie"].id)
        total = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            types=TYPES_CP,
            user_id=users["salarie"].id,
        )
        assert total == 5

    def test_statut_filtre(self, db_session, users, parametrage):
        _ajouter_conge(db_session, users["salarie"].id, statut="en_attente_rh")
        valide = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            types=TYPES_CP,
            user_id=users["salarie"].id,
        )
        en_attente = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUTS_EN_ATTENTE,
            types=TYPES_CP,
            user_id=users["salarie"].id,
        )
        assert valide == 0
        assert en_attente == 5

    def test_conge_id_exclu(self, db_session, users, parametrage):
        c = _ajouter_conge(db_session, users["salarie"].id)
        total = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            types=TYPES_CP,
            user_id=users["salarie"].id,
            conge_id_exclu=c.id,
        )
        assert total == 0

    def test_group_by_user(self, db_session, users, parametrage):
        _ajouter_conge(db_session, users["salarie"].id)
        _ajouter_conge(db_session, users["salarie_sans_resp"].id, nb_jours_ouvrables=3)
        res = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            types=TYPES_CP,
            user_ids=[users["salarie"].id, users["salarie_sans_resp"].id],
            group_by="user",
        )
        assert res[users["salarie"].id] == 5
        assert res[users["salarie_sans_resp"].id] == 3

    def test_group_by_user_type(self, db_session, users, parametrage):
        _ajouter_conge(db_session, users["salarie"].id, type_conge="CP")
        _ajouter_conge(
            db_session, users["salarie"].id, type_conge="Maladie",
            date_debut=date(2026, 7, 1), date_fin=date(2026, 7, 2), nb_jours_ouvrables=2,
        )
        res = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            user_ids=[users["salarie"].id],
            group_by="user_type",
        )
        assert res[(users["salarie"].id, "CP")] == 5
        assert res[(users["salarie"].id, "Maladie")] == 2

    def test_demi_journee_non_tronquee(self, db_session, users, parametrage):
        _ajouter_conge(
            db_session, users["salarie"].id, nb_jours_ouvrables=4.5,
            date_debut=date(2026, 6, 1), date_fin=date(2026, 6, 5),
        )
        total = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice,
            date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE,
            types=TYPES_CP,
            user_id=users["salarie"].id,
        )
        assert total == 4.5


class TestProrataFrontiere:
    """Décompte au prorata des congés à cheval sur une borne d'exercice (R1)."""

    def test_conge_a_cheval_reparti_entre_exercices(self, db_session, users):
        # Congé 30/12/2026 → 03/01/2027 : doit être réparti entre les deux exercices.
        debut, fin = date(2026, 12, 30), date(2027, 1, 3)
        njo = compter_jours_ouvrables_avec_demi(debut, fin)
        _ajouter_conge(
            db_session, users["salarie"].id,
            date_debut=debut, date_fin=fin, nb_jours_ouvrables=njo, type_conge="CP",
        )

        part_2026 = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=date(2026, 1, 1), date_fin_max=date(2026, 12, 31),
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )
        part_2027 = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=date(2027, 1, 1), date_fin_max=date(2027, 12, 31),
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )

        # Aucun jour perdu ni compté deux fois : la somme des parts == total.
        assert part_2026 > 0
        assert part_2027 > 0
        assert part_2026 + part_2027 == njo
        # Chaque part correspond aux jours ouvrables réellement dans la fenêtre.
        assert part_2026 == compter_jours_ouvrables_avec_demi(date(2026, 12, 30), date(2026, 12, 31))
        assert part_2027 == compter_jours_ouvrables_avec_demi(date(2027, 1, 1), date(2027, 1, 3))

    def test_demi_journee_de_bordure_dans_la_bonne_part(self, db_session, users):
        # Congé qui commence l'après-midi du 31/12/2026 et finit le 02/01/2027.
        debut, fin = date(2026, 12, 31), date(2027, 1, 2)
        njo = compter_jours_ouvrables_avec_demi(debut, fin, demi_debut="apres_midi")
        _ajouter_conge(
            db_session, users["salarie"].id,
            date_debut=debut, date_fin=fin, nb_jours_ouvrables=njo,
            demi_journee_debut="apres_midi", type_conge="CP",
        )

        part_2026 = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=date(2026, 1, 1), date_fin_max=date(2026, 12, 31),
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )
        part_2027 = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=date(2027, 1, 1), date_fin_max=date(2027, 12, 31),
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )
        # La demi-journée du 31/12 tombe côté 2026 : 0,5 j en 2026.
        assert part_2026 == 0.5
        assert part_2026 + part_2027 == njo

    def test_conge_entierement_contenu_inchange(self, db_session, users, parametrage):
        # Non-régression : un congé entièrement dans la fenêtre est compté tel quel.
        _ajouter_conge(db_session, users["salarie"].id, nb_jours_ouvrables=5)
        total = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=parametrage.debut_exercice, date_fin_max=parametrage.fin_exercice,
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )
        assert total == 5

    def test_export_as_of_compte_partie_ecoulee(self, db_session, users):
        # Export « à une date » : un congé en cours est compté pour sa part écoulée.
        debut, fin = date(2026, 6, 1), date(2026, 6, 30)  # lun → mar
        njo = compter_jours_ouvrables_avec_demi(debut, fin)
        _ajouter_conge(
            db_session, users["salarie"].id,
            date_debut=debut, date_fin=fin, nb_jours_ouvrables=njo, type_conge="CP",
        )
        # Arrêté au 15/06 : seule la part 01→15 doit compter.
        consomme = somme_consommation(
            colonne=Conge.nb_jours_ouvrables,
            date_debut_min=date(2026, 1, 1), date_fin_max=date(2026, 6, 15),
            statuts=STATUT_VALIDE, types=TYPES_CP, user_id=users["salarie"].id,
        )
        assert consomme == compter_jours_ouvrables_avec_demi(date(2026, 6, 1), date(2026, 6, 15))
        assert 0 < consomme < njo


class TestCoherenceInterEcrans:
    """0 écart entre l'écran salarié, l'export comptable et l'intéressement (NFR9)."""

    def _consomme_export_compta_cp(self, parametrage, user):
        buffer = export_compta_cp_rtt_xlsx(parametrage, parametrage.fin_exercice)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Synthèse CP"]
        nom_complet = f"{user.prenom} {user.nom}"
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == nom_complet:
                return row[2]  # colonne "Consommé (jours)"
        return None

    def test_cp_coherent_solde_export_interessement(self, db_session, users, parametrage, allocations):
        user = users["salarie"]
        _ajouter_conge(db_session, user.id, nb_jours_ouvrables=5, type_conge="CP")
        _ajouter_conge(
            db_session, user.id, nb_jours_ouvrables=2, type_conge="Anciennete",
            date_debut=date(2026, 3, 2), date_fin=date(2026, 3, 3),
        )

        solde_consomme = calculer_jours_cps_consommes(user.id)
        export_consomme = self._consomme_export_compta_cp(parametrage, user)

        from models.interessement_periode import InteressementPeriode
        periode = InteressementPeriode(
            libelle="2026", date_debut=parametrage.debut_exercice,
            date_fin=parametrage.fin_exercice, base_points=100, plancher_points=0, actif=True,
        )
        db_session.session.add(periode)
        db_session.session.commit()
        res = [r for r in calculer_interessement(periode) if r.user_id == user.id][0]
        interessement_cp = sum(d.jours for d in res.details if d.type_absence in ("CP", "Anciennete"))

        assert solde_consomme == 7
        assert export_consomme == 7
        assert interessement_cp == 7

    def test_rtt_coherent_solde_export(self, db_session, users, parametrage, allocations):
        user = users["salarie"]
        _ajouter_conge(
            db_session, user.id, type_conge="RTT", nb_heures_rtt=7, nb_jours_ouvrables=0,
            date_debut=date(2026, 5, 4), date_fin=date(2026, 5, 4),
        )

        solde_rtt = calculer_heures_rtt_consommes(user.id)

        buffer = export_compta_cp_rtt_xlsx(parametrage, parametrage.fin_exercice)
        wb = openpyxl.load_workbook(buffer)
        ws = wb["Synthèse RTT"]
        nom_complet = f"{user.prenom} {user.nom}"
        export_rtt = None
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == nom_complet:
                export_rtt = row[2]
                break

        assert solde_rtt == 7
        assert export_rtt == 7
